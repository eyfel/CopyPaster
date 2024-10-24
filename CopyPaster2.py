import re
from pynput.mouse import Listener as MouseListener
import pyautogui
import pyperclip
import pytesseract
from PIL import Image, ImageEnhance
from tkinter import Tk, Label
import threading
from openpyxl import load_workbook

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

workbook = load_workbook('data.xlsx')
sheet = workbook.active

copy_list = [{"copy_id": row[7], "copy_value": row[6]} for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row - 1, values_only=True)]

copy_id_set = set(item["copy_id"] for item in copy_list)

copy_id_count = {}
for item in copy_list:
    copy_id = item["copy_id"]
    if copy_id in copy_id_count:
        copy_id_count[copy_id] += 1
    else:
        copy_id_count[copy_id] = 1

for copy_id, count in copy_id_count.items():
    if count > 1:
        print(f"Duplicate copy_id {copy_id} found {count} times in the table.")

def update_gui_with_copy_id(copy_id):
    window.title("Copied ID")
    label.config(text=f"{copy_id}")
    window.update_idletasks()
    window.update()
    window.after(200, lambda: label.config(bg="green"))
    window.after(400, lambda: label.config(bg="blue"))
    window.after(600, lambda: label.config(bg="green"))
    window.after(800, lambda: label.config(bg="blue"))
    window.after(1000, lambda: label.config(bg="SystemButtonFace"))

def keep_window_always_on_top():
    window.attributes('-topmost', True)

def start_gui():
    global window, label
    window = Tk()
    window.title("Copied ID")
    window_width = 800
    window_height = 110
    screen_width = window.winfo_screenwidth()
    x_coordinate = (screen_width / 2) - (window_width / 2)
    y_coordinate = 0 
    window.geometry(f"{window_width}x{window_height}+{int(x_coordinate)}+{int(y_coordinate)}")
    label = Label(window, font=("Arial", 48, "bold"), foreground="red", text="No ID copied yet")
    label.pack(expand=True)
    window.attributes('-topmost', True)
    window.mainloop()

gui_thread = threading.Thread(target=start_gui, daemon=True)
gui_thread.start()

def find_longest_match(text):
    matched_items = []
    for item in copy_list:
        if item['copy_id'] in text:
            matched_items.append(item)
    
    if matched_items:
        longest_match = max(matched_items, key=lambda item: len(item["copy_id"]))
        return longest_match
    return None

def read_and_copy(x, y, button, pressed):
    if pressed:
        keep_window_always_on_top()

    if button.name == "left" and pressed:
        region_to_capture = (460, 110, 480, 120)
        screenshot = pyautogui.screenshot(region=region_to_capture)
        screenshot.save("screenshot.png")

        image = Image.open("screenshot.png")
        gray_image = image.convert('L')
        enhancer = ImageEnhance.Contrast(gray_image)
        enhanced_image = enhancer.enhance(2)
        text = pytesseract.image_to_string(enhanced_image)
        text = re.sub(r'[.]', '-', text)
        print(text)

        longest_match = find_longest_match(text)
        if longest_match:
            pyperclip.copy(longest_match["copy_value"])
            update_gui_with_copy_id(longest_match["copy_id"])

with MouseListener(on_click=read_and_copy) as listener:
    listener.join()
